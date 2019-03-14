# -*- coding: UTF-8 -*-
#!/usr/bin/python
import os
import itchat
from itchat.content import *
import urllib
import urllib2
import requests
import random
import openpyxl
from openpyxl import Workbook
import xlsxwriter
import re
import time
from bs4 import BeautifulSoup
import sys
reload(sys) 
sys.setdefaultencoding('utf8')


def get_book_url(url):
    print url
    my_headers=[
"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14",  
"Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)"
]  
    randdom_header=random.choice(my_headers)
    req = urllib2.Request(url)
    req.add_header("User-Agent",randdom_header)
    req.add_header("GET", url)
    response = urllib2.urlopen(req)
    print response.getcode()
    html = response.read()
    bookInfoUrl = []
    weiyunShare = []
    soup = BeautifulSoup(html, 'html.parser')
    for item in soup.find_all('a'):
        if item['href'].startswith('/index.php/bookInfo') and '8822.html' not in item['href'] and '1503.html' not in item['href']:
            #print item['href']
            bookInfoUrl.append('http://www.ireadweek.com' + item['href'])
            #weiyunShare.append(get_weiyun_url('http://www.ireadweek.com' + item['href']))
            shareUrl = get_weiyun_url('http://www.ireadweek.com' + item['href'])
            if 'baidu' in shareUrl:
                with open('baidu.txt','a') as fp:
                    fp.write(shareUrl + '\n')
                    fp.close()
            #time.sleep(10)
    '''
    print len(bookInfoUrl)
    print 'get book name'
    bookInfoName = []
    for tag in soup.find_all('div', 'hanghang-list-name'):
        print tag.string
        bookInfoName.append(tag.string)
	if '书籍' in bookInfoName:
	    bookInfoName.remove('书籍')
    if '捐助' in bookInfoName:
        bookInfoName.remove('捐助')
    if '下载方式及帮助' in bookInfoName:
        bookInfoName.remove('下载方式及帮助')
    print 'get authors'
    bookInfoAuthor = []
    for tag in soup.find_all('div', 'hanghang-list-zuozhe'):
        print tag.string
        bookInfoAuthor.append(tag.string)
    if '作者' in bookInfoAuthor:
        bookInfoAuthor.remove('作者')
    if '行行' in bookInfoAuthor:
        bookInfoAuthor.remove('行行')
    if '小编：行行' in bookInfoAuthor:
        bookInfoAuthor.remove('小编：行行')
    print len(bookInfoName)
    print len(bookInfoAuthor)
    print len(weiyunShare)
    
    writeExcel(bookInfoName, bookInfoAuthor, bookInfoUrl, weiyunShare)
    print '-----------------------------------------------------'
    '''

def get_weiyun_url(url):
    #print 'get_weiyun_url'
    my_headers=[
"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.153 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.1; WOW64; rv:30.0) Gecko/20100101 Firefox/30.0",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/537.75.14",  
"Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Win64; x64; Trident/6.0)"
]  
    randdom_header=random.choice(my_headers)
    req = urllib2.Request(url)
    req.add_header("User-Agent",randdom_header)
    req.add_header("GET", url)
    response = urllib2.urlopen(req)
    print response.getcode()
    html = response.read()    
    soup = BeautifulSoup(html, 'html.parser')
    for item in soup.find_all('a', 'downloads'):
        print item['href']
        return item['href']
    return ''
    
def createExcel():
    print 'createExcel'
    headings = ['书名', '作者', '书籍详细信息', '云盘地址']
    wb = Workbook()
    ws = wb.create_sheet('ireadweek')
    ws.append(headings)
    wb.save('ireadweek.xlsx')

    
    
def writeExcel(bookname = [], authors = [], urls = [], weiyun = []):
    excel = openpyxl.load_workbook('ireadweek.xlsx')
    table = excel.get_sheet_by_name('ireadweek')
    nrows = table.max_row
    ncols = table.max_column
    #print 'nrows:' + str(nrows)
    #print 'ncols:' + str(ncols)
    for i in range(len(bookname)):
        if authors[i] == None:
            authors[i] = '不详'
        print str(bookname[i]) + ';' + str(authors[i]) + ';' + str(urls[i]) + ';' + str(weiyun[i])
        values = []
        values.append(bookname[i].replace(' ', ''))
        values.append(authors[i].decode('utf-8'))
        values.append(urls[i])
        values.append(weiyun[i])
        table.append(values)
    excel.save('ireadweek.xlsx')    

if __name__ == '__main__':
    if not os.path.exists(os.getcwd() + '/ireadweek.xlsx'):
        print 'create excel'
        createExcel()
    for i in range(1, 3):
        get_book_url(u'http://www.ireadweek.com/index.php/index/%d.html' % i)

	
	

